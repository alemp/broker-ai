"""Phase 5 — parse, validate, and apply CSV/Excel client imports."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import unicodedata
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

from email_validator import EmailNotValidError, validate_email
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ai_copilot_api.db.enums import ClientKind, CrmEntityType, IngestionSource, InsuredRelation
from ai_copilot_api.db.models import (
    Client,
    ClientHeldProduct,
    ClientImportBatch,
    ClientLineOfBusiness,
    InsuredPerson,
    LineOfBusiness,
    Product,
    User,
)
from ai_copilot_api.domain.client_profile import merge_profile_dict, parse_profile
from ai_copilot_api.domain.crm_audit import record_entity_snapshot_create, record_field_updates

_MAX_FILE_BYTES = 15 * 1024 * 1024
_MAX_ROWS = 10_000
_PREVIEW_ROWS = 50

RawImportRow = dict[str, str | None]
HeldSegmentTuple = tuple[str | None, str | None, str | None, date | None, date | None]

_HEADER_ALIASES: dict[str, str] = {
    # full_name
    "nome": "full_name",
    "name": "full_name",
    "nome_completo": "full_name",
    "nome_do_cliente": "full_name",
    "nome_cliente": "full_name",
    # email
    "e_mail": "email",
    "correio": "email",
    "mail": "email",
    "correio_eletronico": "email",
    "endereco_eletronico": "email",
    "endereco_de_email": "email",
    "endereco_email": "email",
    # external_id
    "id_externo": "external_id",
    "externalid": "external_id",
    "identificador_externo": "external_id",
    "codigo_externo": "external_id",
    "ref_externa": "external_id",
    "referencia_externa": "external_id",
    # phone
    "telefone": "phone",
    "tel": "phone",
    "telemovel": "phone",
    "celular": "phone",
    "contacto_telefonico": "phone",
    "contato_telefonico": "phone",
    # notes
    "observacoes": "notes",
    "anotacoes": "notes",
    "comentarios": "notes",
    # owner_email
    "corretor_email": "owner_email",
    "owner_email": "owner_email",
    "email_do_corretor": "owner_email",
    "email_corretor": "owner_email",
    "email_responsavel": "owner_email",
    "responsavel_email": "owner_email",
    # client_kind
    "tipo_cliente": "client_kind",
    "tipo": "client_kind",
    "tipo_de_cliente": "client_kind",
    "natureza_cliente": "client_kind",
    # company
    "razao_social": "company_legal_name",
    "nome_empresa": "company_legal_name",
    "denominacao_social": "company_legal_name",
    "cnpj": "company_tax_id",
    "nif": "company_tax_id",
    "contribuinte": "company_tax_id",
    "matricula_fiscal": "company_tax_id",
    "codigo_fiscal": "company_tax_id",
    "numero_fiscal": "company_tax_id",
    # marketing
    "opt_in_marketing": "marketing_opt_in",
    "aceita_marketing": "marketing_opt_in",
    "consentimento_marketing": "marketing_opt_in",
    "autoriza_marketing": "marketing_opt_in",
    "marketing_opt_in": "marketing_opt_in",
    "canal_preferido": "preferred_marketing_channel",
    "canal_marketing": "preferred_marketing_channel",
    "canal_de_comunicacao": "preferred_marketing_channel",
    "canal_comunicacao": "preferred_marketing_channel",
    "preferred_marketing_channel": "preferred_marketing_channel",
    # LOB / held / profile / insured (canonical + PT)
    "linhas_negocio": "lob_codes",
    "linhas_de_negocio": "lob_codes",
    "linha_negocio": "lob_codes",
    "codigos_lob": "lob_codes",
    "codigo_lob": "lob_codes",
    "lob_codes": "lob_codes",
    "lobs": "lob_codes",
    "produtos_detidos": "held_products",
    "produtos_em_carteira": "held_products",
    "carteira_seguros": "held_products",
    "held_products": "held_products",
    "perfil_json": "profile_json",
    "perfil_em_json": "profile_json",
    "dados_perfil": "profile_json",
    "dados_perfil_json": "profile_json",
    "profile_json": "profile_json",
    "segurados_json": "insured_persons_json",
    "lista_segurados_json": "insured_persons_json",
    "segurados_em_json": "insured_persons_json",
    "insured_persons_json": "insured_persons_json",
}


def _fold_accents(s: str) -> str:
    """ASCII-fold for header matching (accents removed, e.g. razão → razao)."""
    decomposed = unicodedata.normalize("NFD", s)
    return "".join(c for c in decomposed if unicodedata.category(c) != "Mn")


def normalize_header(raw: str) -> str:
    s = raw.strip().lower().replace(" ", "_").replace("-", "_")
    while "__" in s:
        s = s.replace("__", "_")
    folded = _fold_accents(s)
    return _HEADER_ALIASES.get(folded, folded)


def sha256_hex(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def detect_format(filename: str | None, content: bytes) -> str:
    fn = (filename or "").lower()
    if fn.endswith(".xlsx"):
        return "xlsx"
    if fn.endswith(".csv"):
        return "csv"
    if len(content) >= 2 and content[:2] == b"PK":
        return "xlsx"
    return "csv"


def read_upload_limited(content: bytes) -> tuple[bytes | None, str | None]:
    if len(content) > _MAX_FILE_BYTES:
        mb = _MAX_FILE_BYTES // (1024 * 1024)
        return None, f"A planilha excede o tamanho máximo permitido ({mb} MB)."
    return content, None


def _pt_email_validation_error(detail: str) -> str:
    """Brazilian Portuguese message for email-validator errors (library text is English)."""
    d = detail.lower()
    if "full-width" in d or "small commercial at" in d:
        return "Use o símbolo @ normal (ASCII), não uma variante Unicode."
    if "must have an @-sign" in d or "an email address must have an @-sign" in d:
        return "O e-mail deve incluir o símbolo @."
    if "something before the @" in d or "before the @-sign" in d:
        return "É necessário haver texto antes do @."
    if "after the @-sign" in d or "domain part" in d:
        return "A parte depois do @ (domínio) não é válida."
    if "not valid" in d or "invalid" in d:
        return "O formato do e-mail não é válido."
    return "O endereço de e-mail não é válido."


def _pt_json_decode_error(campo_pt: str, e: json.JSONDecodeError) -> str:
    return (
        f"{campo_pt}: sintaxe JSON inválida "
        f"(linha {e.lineno}, coluna {e.colno})."
    )


def _cell_to_str(val: object | None) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date().isoformat() if val.tzinfo is None else val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    return s if s else None


def parse_spreadsheet(
    content: bytes,
    source_format: str,
) -> tuple[list[RawImportRow], str | None]:
    """Returns raw rows as dicts (canonical header keys) or a fatal error message."""
    if source_format == "csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = content.decode("latin-1")
            except UnicodeDecodeError:
                return [], (
                    "O conteúdo da planilha não é texto válido em UTF-8 nem em Latin-1."
                )
        f = io.StringIO(text)
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return [], "O CSV não tem linha de cabeçalho."
        rows: list[RawImportRow] = []
        for i, row_dict in enumerate(reader):
            if i >= _MAX_ROWS:
                return [], f"Número de linhas acima do máximo permitido ({_MAX_ROWS})."
            out: RawImportRow = {}
            for k, v in row_dict.items():
                if k is None:
                    continue
                nk = normalize_header(k)
                if not nk:
                    continue
                out[nk] = v.strip() if isinstance(v, str) else _cell_to_str(v)
            rows.append(out)
        return rows, None

    try:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        return [], f"Planilha Excel inválida ou corrompida: {e!s}"
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header_row = next(it, None)
        if not header_row:
            return [], (
                "A planilha Excel está vazia ou a primeira linha não contém cabeçalhos."
            )
        headers: list[str] = []
        for h in header_row:
            hs = _cell_to_str(h) or ""
            headers.append(normalize_header(hs) if hs else "")
        rows: list[RawImportRow] = []
        data_index = 0
        for _sheet_row_idx, row in enumerate(it, start=2):
            cells = list(row) if row else []
            if not any(_cell_to_str(c) for c in cells):
                continue
            if data_index >= _MAX_ROWS:
                return [], f"Número de linhas acima do máximo permitido ({_MAX_ROWS})."
            data_index += 1
            out: RawImportRow = {}
            for j, key in enumerate(headers):
                if not key:
                    continue
                val = cells[j] if j < len(cells) else None
                out[key] = _cell_to_str(val)
            rows.append(out)
        return rows, None
    finally:
        wb.close()


def _validate_email_or_none(s: str | None) -> tuple[str | None, str | None]:
    if s is None or not s.strip():
        return None, None
    try:
        norm = validate_email(s.strip(), check_deliverability=False).normalized
        return norm.lower(), None
    except EmailNotValidError as e:
        return None, str(e)


def _norm_external_id(s: str | None) -> str | None:
    if s is None:
        return None
    t = s.strip()
    return t if t else None


def _parse_bool(s: str | None) -> bool | None:
    if s is None or not s.strip():
        return None
    v = s.strip().lower()
    if v in ("1", "true", "yes", "y", "sim", "s"):
        return True
    if v in ("0", "false", "no", "n", "nao", "não"):
        return False
    return None


def _parse_client_kind(s: str | None) -> ClientKind | None:
    if s is None or not s.strip():
        return None
    v = s.strip().upper().replace(" ", "_")
    if v in ("COMPANY", "EMPRESA", "PJ", "CORPORATE"):
        return ClientKind.COMPANY
    if v in ("INDIVIDUAL", "PF", "PESSOA_FISICA", "PESSOA_FÍSICA", "FISICA", "FÍSICA"):
        return ClientKind.INDIVIDUAL
    return None


def _parse_date(s: str | None) -> date | None:
    if s is None or not str(s).strip():
        return None
    t = str(s).strip()[:10]
    try:
        return date.fromisoformat(t)
    except ValueError:
        return None


def _parse_insured_relation(s: str) -> InsuredRelation:
    u = s.strip().upper().replace(" ", "_")
    for m in InsuredRelation:
        if m.name == u or m.value == u:
            return m
    return InsuredRelation.HOLDER


@dataclass
class ValidatedImportRow:
    row_number: int
    external_id: str | None
    email: str | None
    full_name: str
    phone: str | None
    notes: str | None
    owner_email: str | None
    owner_id: uuid.UUID | None
    client_kind: ClientKind
    company_legal_name: str | None
    company_tax_id: str | None
    marketing_opt_in: bool | None
    preferred_marketing_channel: str | None
    lob_codes: list[str]
    held_segments: list[HeldSegmentTuple]
    profile_patch: dict[str, Any]
    insured_specs: list[tuple[str, InsuredRelation]]


@dataclass
class ImportValidationResult:
    rows: list[ValidatedImportRow] = field(default_factory=list)
    errors: list[tuple[int, str]] = field(default_factory=list)


def validate_rows_structural(raw_rows: list[RawImportRow]) -> ImportValidationResult:
    result = ImportValidationResult()
    seen_email: set[str] = set()
    seen_ext: set[str] = set()
    for idx, r in enumerate(raw_rows, start=2):
        ext = _norm_external_id(r.get("external_id"))
        em_raw = r.get("email")
        em, em_err = _validate_email_or_none(em_raw.strip() if em_raw else None)
        if em_err:
            result.errors.append((idx, _pt_email_validation_error(em_err)))
            continue
        if ext is None and em is None:
            result.errors.append(
                (
                    idx,
                    "Cada linha deve incluir pelo menos o e-mail ou o ID externo "
                    "(identificação para criar ou atualizar o cliente).",
                ),
            )
            continue
        if em and em in seen_email:
            result.errors.append((idx, f"E-mail duplicado na planilha: {em}"))
            continue
        if ext and ext in seen_ext:
            result.errors.append((idx, f"ID externo duplicado na planilha: {ext}"))
            continue
        if em:
            seen_email.add(em)
        if ext:
            seen_ext.add(ext)

        full_name = (r.get("full_name") or "").strip()
        if not full_name:
            result.errors.append((idx, "O nome completo é obrigatório."))
            continue

        ck = _parse_client_kind(r.get("client_kind")) or ClientKind.INDIVIDUAL
        company_legal = (r.get("company_legal_name") or "").strip() or None
        company_tax = (r.get("company_tax_id") or "").strip() or None
        if ck == ClientKind.COMPANY and not company_legal:
            result.errors.append(
                (
                    idx,
                    "Para clientes do tipo empresa, a razão social "
                    "(nome da empresa) é obrigatória.",
                ),
            )
            continue

        lob_raw = r.get("lob_codes") or ""
        lob_codes: list[str] = []
        if lob_raw and lob_raw.strip():
            for part in re.split(r"[;,]", lob_raw):
                p = part.strip()
                if p:
                    lob_codes.append(p)

        held_segments: list[HeldSegmentTuple] = []
        hp = r.get("held_products")
        if hp and hp.strip():
            for seg in hp.split(";"):
                seg = seg.strip()
                if not seg:
                    continue
                parts = [x.strip() or None for x in seg.split("|")]
                while len(parts) < 5:
                    parts.append(None)
                pn, ins, st, ed_s, end_s = parts[:5]
                held_segments.append(
                    (pn, ins, st, _parse_date(ed_s), _parse_date(end_s)),
                )

        profile_patch: dict[str, Any] = {}
        pj = r.get("profile_json")
        if pj and pj.strip():
            try:
                loaded = json.loads(pj)
            except json.JSONDecodeError as e:
                result.errors.append(
                    (idx, _pt_json_decode_error("Campo Perfil (profile_json)", e)),
                )
                continue
            if not isinstance(loaded, dict):
                result.errors.append(
                    (
                        idx,
                        "O campo Perfil (profile_json) precisa ser um objeto JSON "
                        "(chaves entre { }).",
                    ),
                )
                continue
            profile_patch = loaded

        insured_specs: list[tuple[str, InsuredRelation]] = []
        ij = r.get("insured_persons_json")
        insured_err: str | None = None
        if ij and ij.strip():
            try:
                arr = json.loads(ij)
            except json.JSONDecodeError as e:
                result.errors.append(
                    (idx, _pt_json_decode_error("Campo Segurados (insured_persons_json)", e)),
                )
                continue
            if not isinstance(arr, list):
                result.errors.append(
                    (
                        idx,
                        "O campo Segurados (insured_persons_json) precisa ser uma lista JSON "
                        "(entre [ ]).",
                    ),
                )
                continue
            for item in arr:
                if not isinstance(item, dict):
                    insured_err = "Cada segurado na lista precisa ser um objeto JSON { … }."
                    break
                iname = (item.get("full_name") or item.get("nome") or "").strip()
                if not iname:
                    insured_err = "Cada segurado deve ter nome completo (full_name ou nome)."
                    break
                rel_raw = (item.get("relation") or item.get("relacao") or "HOLDER").strip()
                rel = _parse_insured_relation(rel_raw)
                insured_specs.append((iname, rel))
            if insured_err:
                result.errors.append((idx, insured_err))
                continue

        mo = _parse_bool(r.get("marketing_opt_in"))
        oe = (r.get("owner_email") or "").strip().lower() or None
        if oe:
            oe_val, oe_err = _validate_email_or_none(oe)
            if oe_err:
                result.errors.append(
                    (idx, "E-mail do corretor: " + _pt_email_validation_error(oe_err)),
                )
                continue
            oe = oe_val

        result.rows.append(
            ValidatedImportRow(
                row_number=idx,
                external_id=ext,
                email=em,
                full_name=full_name,
                phone=(r.get("phone") or "").strip() or None,
                notes=(r.get("notes") or "").strip() or None,
                owner_email=oe,
                owner_id=None,
                client_kind=ck,
                company_legal_name=company_legal,
                company_tax_id=company_tax,
                marketing_opt_in=mo,
                preferred_marketing_channel=(r.get("preferred_marketing_channel") or "").strip()
                or None,
                lob_codes=lob_codes,
                held_segments=held_segments,
                profile_patch=profile_patch,
                insured_specs=insured_specs,
            ),
        )
    return result


def _email_expr(col: Any) -> Any:
    return func.lower(func.trim(col))


def validate_rows_with_catalog(
    db: Session,
    org_id: uuid.UUID,
    structural: ImportValidationResult,
) -> ImportValidationResult:
    """Resolve owner emails, LOB codes, and product names against org catalog."""
    out = ImportValidationResult(errors=list(structural.errors))
    lob_rows = db.scalars(
        select(LineOfBusiness).where(LineOfBusiness.organization_id == org_id),
    ).all()
    lob_by_code = {row.code.strip().lower(): row for row in lob_rows}

    prod_rows = db.scalars(
        select(Product).where(
            Product.organization_id == org_id,
            Product.active.is_(True),
        ),
    ).all()
    products_by_lname: dict[str, list[Product]] = {}
    for p in prod_rows:
        key = p.name.strip().lower()
        products_by_lname.setdefault(key, []).append(p)

    for row in structural.rows:
        if row.owner_email:
            u = db.scalar(
                select(User).where(
                    User.organization_id == org_id,
                    _email_expr(User.email) == row.owner_email,
                ),
            )
            if u is None:
                out.errors.append(
                    (
                        row.row_number,
                        f"Não existe usuário na corretora com o e-mail {row.owner_email}.",
                    ),
                )
                continue
            row.owner_id = u.id
        unknown_lobs = [c for c in row.lob_codes if c.strip().lower() not in lob_by_code]
        if unknown_lobs:
            codes = ", ".join(unknown_lobs)
            out.errors.append(
                (
                    row.row_number,
                    f"Código(s) de linha de negócio desconhecido(s): {codes}. "
                    "Cadastre-os no catálogo ou corrija a planilha.",
                ),
            )
            continue
        held_ok = True
        for seg in row.held_segments:
            pname = (seg[0] or "").strip()
            if not pname:
                out.errors.append(
                    (
                        row.row_number,
                        "Em produtos na carteira, falta o nome do produto "
                        "(primeiro campo antes do | em cada segmento).",
                    ),
                )
                held_ok = False
                break
            matches = products_by_lname.get(pname.lower(), [])
            if len(matches) != 1:
                n = len(matches)
                if n == 0:
                    prod_msg = (
                        f'O produto "{pname}" não existe no catálogo ativo ou o nome não confere.'
                    )
                else:
                    prod_msg = (
                        f'O nome "{pname}" corresponde a {n} produtos ativos; '
                        "é necessária exatamente uma correspondência."
                    )
                out.errors.append((row.row_number, prod_msg))
                held_ok = False
                break
        if not held_ok:
            continue

        if row.profile_patch:
            try:
                parse_profile(merge_profile_dict({}, row.profile_patch))
            except ValidationError:
                out.errors.append(
                    (
                        row.row_number,
                        "O perfil (profile_json) contém campos ou tipos inválidos em relação ao "
                        "modelo de dados do CRM.",
                    ),
                )
                continue
            except Exception:
                out.errors.append(
                    (
                        row.row_number,
                        "Não foi possível validar o perfil (profile_json). Verifique a estrutura.",
                    ),
                )
                continue

        out.rows.append(row)
    return out


def _client_audit_dict(row: Client) -> dict[str, Any]:
    return {
        "full_name": row.full_name,
        "email": row.email,
        "phone": row.phone,
        "external_id": row.external_id,
        "notes": row.notes,
        "owner_id": row.owner_id,
        "client_kind": row.client_kind.value,
        "company_legal_name": row.company_legal_name,
        "company_tax_id": row.company_tax_id,
        "marketing_opt_in": row.marketing_opt_in,
        "preferred_marketing_channel": row.preferred_marketing_channel,
    }


def _find_existing_client(
    db: Session,
    org_id: uuid.UUID,
    external_id: str | None,
    email: str | None,
) -> Client | None:
    if external_id:
        hit = db.scalar(
            select(Client).where(
                Client.organization_id == org_id,
                Client.external_id == external_id,
            ),
        )
        if hit:
            return hit
    if email:
        return db.scalar(
            select(Client).where(
                Client.organization_id == org_id,
                _email_expr(Client.email) == email.lower(),
            ),
        )
    return None


def apply_import(
    db: Session,
    *,
    organization_id: uuid.UUID,
    actor_user_id: uuid.UUID,
    filename: str,
    file_sha256: str,
    source_format: str,
    ingestion: IngestionSource,
    rows: list[ValidatedImportRow],
) -> ClientImportBatch:
    lob_rows = db.scalars(
        select(LineOfBusiness).where(LineOfBusiness.organization_id == organization_id),
    ).all()
    lob_by_code = {row.code.strip().lower(): row for row in lob_rows}

    prod_rows = db.scalars(
        select(Product).where(
            Product.organization_id == organization_id,
            Product.active.is_(True),
        ),
    ).all()
    products_by_lname: dict[str, list[Product]] = {}
    for p in prod_rows:
        key = p.name.strip().lower()
        products_by_lname.setdefault(key, []).append(p)

    inserted = 0
    updated = 0

    for row in rows:
        existing = _find_existing_client(db, organization_id, row.external_id, row.email)
        if existing is None:
            c = Client(
                organization_id=organization_id,
                full_name=row.full_name,
                email=row.email,
                phone=row.phone,
                external_id=row.external_id,
                notes=row.notes,
                owner_id=row.owner_id,
                client_kind=row.client_kind,
                company_legal_name=row.company_legal_name,
                company_tax_id=row.company_tax_id,
                marketing_opt_in=True if row.marketing_opt_in is None else row.marketing_opt_in,
                preferred_marketing_channel=row.preferred_marketing_channel,
            )
            db.add(c)
            db.flush()
            record_entity_snapshot_create(
                db,
                organization_id=organization_id,
                actor_user_id=actor_user_id,
                entity_type=CrmEntityType.CLIENT,
                entity_id=c.id,
                snapshot={k: v for k, v in _client_audit_dict(c).items() if v is not None},
            )
            client = c
            inserted += 1
        else:
            client = existing
            before = _client_audit_dict(client)
            updates: dict[str, Any] = {"full_name": row.full_name, "client_kind": row.client_kind}
            if row.email is not None:
                updates["email"] = row.email
            if row.phone is not None:
                updates["phone"] = row.phone
            if row.external_id is not None:
                updates["external_id"] = row.external_id
            if row.notes is not None:
                updates["notes"] = row.notes
            if row.owner_id is not None:
                updates["owner_id"] = row.owner_id
            if row.company_legal_name is not None:
                updates["company_legal_name"] = row.company_legal_name
            if row.company_tax_id is not None:
                updates["company_tax_id"] = row.company_tax_id
            if row.marketing_opt_in is not None:
                updates["marketing_opt_in"] = row.marketing_opt_in
            if row.preferred_marketing_channel is not None:
                updates["preferred_marketing_channel"] = row.preferred_marketing_channel
            for k, v in updates.items():
                setattr(client, k, v)
            db.flush()
            audit_updates: dict[str, Any] = {}
            for k, v in updates.items():
                if isinstance(v, ClientKind):
                    audit_updates[k] = v.value
                else:
                    audit_updates[k] = v
            record_field_updates(
                db,
                organization_id=organization_id,
                actor_user_id=actor_user_id,
                entity_type=CrmEntityType.CLIENT,
                entity_id=client.id,
                before=before,
                updates={
                    k: audit_updates[k] for k in audit_updates if before.get(k) != audit_updates[k]
                },
            )
            updated += 1

        for code in row.lob_codes:
            lob = lob_by_code[code.strip().lower()]
            exists = db.scalar(
                select(ClientLineOfBusiness).where(
                    ClientLineOfBusiness.client_id == client.id,
                    ClientLineOfBusiness.line_of_business_id == lob.id,
                ),
            )
            if exists is None:
                db.add(
                    ClientLineOfBusiness(
                        client_id=client.id,
                        line_of_business_id=lob.id,
                        ingestion_source=ingestion,
                    ),
                )

        for seg in row.held_segments:
            pname = (seg[0] or "").strip()
            prod = products_by_lname[pname.lower()][0]
            db.add(
                ClientHeldProduct(
                    client_id=client.id,
                    product_id=prod.id,
                    insurer_name=seg[1],
                    policy_status=seg[2],
                    effective_date=seg[3],
                    end_date=seg[4],
                    ingestion_source=ingestion,
                ),
            )

        if row.profile_patch:
            merged = merge_profile_dict(_raw_profile_dict(client), row.profile_patch)
            client.profile_data = merged
            parse_profile(merged)

        for iname, rel in row.insured_specs:
            db.add(
                InsuredPerson(
                    organization_id=organization_id,
                    client_id=client.id,
                    full_name=iname,
                    relation=rel,
                ),
            )

    batch = ClientImportBatch(
        organization_id=organization_id,
        actor_user_id=actor_user_id,
        filename=filename[:255],
        file_sha256=file_sha256,
        source_format=source_format,
        row_count=len(rows),
        inserted_count=inserted,
        updated_count=updated,
        error_count=0,
    )
    db.add(batch)
    db.flush()
    return batch


def _raw_profile_dict(client: Client) -> dict[str, Any]:
    raw = client.profile_data
    if raw is None:
        return {}
    if isinstance(raw, dict):
        return dict(raw)
    return {}
