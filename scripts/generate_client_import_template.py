#!/usr/bin/env python3
"""Generate apps/web/public/templates/importacao-clientes-modelo.xlsx.

Run from repository root:
  cd apps/api && uv run python ../../scripts/generate_client_import_template.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
OUT = REPO / "apps/web/public/templates/importacao-clientes-modelo.xlsx"


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Clientes"

    headers = [
        "Nome completo",
        "E-mail",
        "ID externo",
        "Telefone",
        "Observações",
        "Email do corretor",
        "Tipo de cliente",
        "Razão social",
        "NIF",
        "Consentimento marketing",
        "Canal preferido",
        "Linhas de negócio",
        "Produtos detidos",
        "Perfil JSON",
        "Segurados JSON",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    example = [
        "Maria Exemplo",
        "maria.exemplo@email.com",
        "EXT-0001",
        "+55 11 99999-9999",
        "Exclua esta linha de exemplo antes de importar dados reais ou substitua pelos seus clientes.",
        "",
        "INDIVIDUAL",
        "",
        "",
        "sim",
        "email",
        "",
        "",
        "",
        "",
    ]
    ws.append(example)
    ws.freeze_panes = "A2"

    widths = [22, 28, 14, 16, 40, 28, 14, 28, 14, 16, 22, 24, 48, 20, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wsi = wb.create_sheet("Instrucoes", 1)
    wsi.column_dimensions["A"].width = 28
    wsi.column_dimensions["B"].width = 88
    title_font = Font(bold=True, size=12)
    wsi["A1"] = "Importação de clientes — modelo de planilha Excel"
    wsi["A1"].font = title_font
    wsi.merge_cells("A1:B1")

    rows: list[tuple[str, str]] = [
        ("", ""),
        (
            "Aba Clientes",
            "A primeira linha são os cabeçalhos — não apague nem renomeie se quiser usar as colunas. "
            "Você pode excluir a linha de exemplo (linha 2) e preencher a partir da linha 2 com seus dados.",
        ),
        (
            "Nome completo",
            "Obrigatório. Também são aceitos: nome, full_name (inglês).",
        ),
        (
            "E-mail",
            "Pelo menos e-mail ou ID externo. Também: correio, e-mail, email.",
        ),
        (
            "ID externo",
            "Código no sistema antigo (único na corretora). Também: id_externo, identificador externo.",
        ),
        ("Telefone", "Opcional. Também: celular, tel."),
        ("Observações", "Opcional. Também: anotações, notes."),
        (
            "Email do corretor",
            "Opcional: usuário responsável na corretora. Também: corretor_email, owner_email.",
        ),
        (
            "Tipo de cliente",
            "INDIVIDUAL (padrão) ou COMPANY / EMPRESA / PJ. Também: tipo_cliente, client_kind.",
        ),
        (
            "Razão social",
            "Obrigatório se for empresa. Também: company_legal_name, nome_empresa.",
        ),
        ("NIF", "Opcional (fiscal). Também: CNPJ, company_tax_id, contribuinte."),
        (
            "Consentimento marketing",
            "Opcional: sim/não, 1/0. Também: marketing_opt_in, aceita_marketing.",
        ),
        (
            "Canal preferido",
            "Canal de contato de marketing. Também: preferred_marketing_channel, canal_marketing.",
        ),
        (
            "Linhas de negócio",
            "Códigos na app, separados por vírgula ou ;. Vazio se não usar. Também: lob_codes.",
        ),
        (
            "Produtos detidos",
            "Carteira: vários produtos separados por ;. Cada um: produto|seguradora|estado|início|fim "
            "(AAAA-MM-DD). Também: held_products.",
        ),
        (
            "Perfil JSON",
            "Avançado: JSON do perfil de seguros. Também: profile_json, perfil_json.",
        ),
        (
            "Segurados JSON",
            "Avançado: JSON com lista de segurados. Também: insured_persons_json, segurados_json.",
        ),
        (
            "",
            "",
        ),
        (
            "Validação",
            "Use Visualizar na aplicação antes de confirmar. Códigos LOB e nomes de produtos são "
            "validados com o catálogo da sua corretora.",
        ),
    ]

    r = 2
    for a, b in rows:
        wsi.cell(row=r, column=1, value=a).alignment = Alignment(wrap_text=True, vertical="top")
        c = wsi.cell(row=r, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
